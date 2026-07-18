#!/usr/bin/env python3
"""
gold_unified_sheet.py — the strain-level decision card (what Adam sees).

Combines the roster (identity_strains) with the silver stat sheets into ONE row per strain.
Silver tables are at isolate grain; we map each isolate to its strain group and aggregate per the
team-owned policy in config/thresholds.yaml (safety = worst_case, competition = best_case).
Candidate/safety status comes from data/reference/species_safety.csv (team-owned; swap in Gwyn's BSL list).
Reserved columns (tissue/mouse/ubiquity) are blank until that data lands.

EVERY threshold and policy here is team-owned and one edit away — see config/thresholds.yaml.
Decisions: docs/decisions/gold_unified_sheet_decisions.md

Run:  python gold_unified_sheet.py
Out:  ../data/gold/gold_unified_sheet.{csv,parquet,xlsx}
"""
import os
import re
from collections import defaultdict
from lib_ids import read_delimited, write_table
from config import CFG
from data_sources import is_enabled

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
REF = os.path.join(ROOT, "data", "reference")
SILVER = os.path.join(ROOT, "data", "silver")
GOLD = os.path.join(ROOT, "data", "gold", "gold_unified_sheet")
SAFETY_REF = os.path.join(REF, "species_safety.csv")

SAFE_AGG = CFG["gold"]["safety_aggregation"]        # worst_case
COMP_AGG = CFG["gold"]["competition_aggregation"]   # best_case

GOLD_COLS = [
    "strain_group", "representative_asma_id", "genus", "species", "n_isolates",
    "is_candidate", "candidate_review", "bsl_level",
    "hemolysis_beta", "hemolysis_concern", "amr_resistance_count_prov", "amr_gene_count",
    "grows_scfm", "scfm_od", "mucin_lift",
    "comp_best_solo_pa", "comp_best_team_pa", "comp_best_partner", "comp_synergy_pa", "comp_n_formulations",
    "tissue", "mouse", "airway_ubiquity", "decision", "decision_reason",
]


def num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def norm_species(s):
    """Drop GTDB alphabetic suffixes so 'Citrobacter_B koseri' matches 'Citrobacter koseri'."""
    return re.sub(r"_[A-Z]+\b", "", s).strip() if s else s


def load_keyed(path, key="asma_id"):
    return {r[key]: r for r in read_delimited(path, ",")}


def load_safety_ref():
    by_species, by_genus = {}, {}
    for r in read_delimited(SAFETY_REF, ","):
        name = r["species"].strip()
        (by_genus if " " not in name else by_species)[name] = r
    return by_species, by_genus


def classify(species, genus, by_species, by_genus):
    """-> (is_candidate, review_status, bsl_level). Species match wins over genus; else 'unreviewed'."""
    ss = by_species.get(norm_species(species)) or by_genus.get(norm_species(genus or ""))
    level = ss["level"] if ss else "unreviewed"
    bsl = (ss["bsl_level"] if ss else "") or ""
    return (level not in ("pathogen", "opportunistic"), level, bsl)


def main():
    os.makedirs(os.path.dirname(GOLD), exist_ok=True)
    strains = list(read_delimited(os.path.join(REF, "identity_strains.csv"), ","))
    grp_members = defaultdict(list)
    for iso in read_delimited(os.path.join(REF, "identity_isolates.csv"), ","):
        grp_members[iso["strain_group"]].append(iso["asma_id"])

    # Only join a source's silver table if that source is enabled in data_sources.yaml.
    # A disabled source -> empty dict -> its columns come out blank (like not-yet-arrived data).
    hemo = load_keyed(os.path.join(SILVER, "silver_hemolysis.csv")) if is_enabled("hemolysis") else {}
    amrm = load_keyed(os.path.join(SILVER, "silver_amr_measured.csv")) if is_enabled("amr_measured") else {}
    amrg = load_keyed(os.path.join(SILVER, "silver_amr_genomic.csv")) if is_enabled("amr_genomic") else {}
    comp = load_keyed(os.path.join(SILVER, "silver_competition.csv")) if is_enabled("competition") else {}
    grow = load_keyed(os.path.join(SILVER, "silver_growth_endpoint.csv")) if is_enabled("growth_endpoint") else {}
    by_species, by_genus = load_safety_ref()

    rows = []
    for s in strains:
        grp = s["strain_group"]
        members = grp_members.get(grp, [s["representative_asma_id"]])
        species, genus = s.get("species") or None, s.get("genus")
        is_cand, review, bsl = classify(species, genus, by_species, by_genus)

        # SAFETY — worst-case across isolates
        betas = [hemo[a]["beta_hemolytic"] for a in members if a in hemo]
        beta = "Y" if "Y" in betas else ("N" if "N" in betas else None)
        res = [int(amrm[a]["resistance_count_prov"]) for a in members
               if a in amrm and amrm[a]["resistance_count_prov"] not in ("", None)]
        genes = [int(amrg[a]["amr_gene_count"]) for a in members
                 if a in amrg and amrg[a]["amr_gene_count"] not in ("", None)]

        # VIABILITY — best-case across isolates
        gcalls = [grow[a]["grows_scfm"] for a in members if a in grow]
        grows = "Y" if "Y" in gcalls else ("N" if "N" in gcalls else None)
        scfm_ods = [num(grow[a]["scfm_od"]) for a in members if a in grow and num(grow[a]["scfm_od"]) is not None]
        lifts = [num(grow[a]["mucin_lift"]) for a in members if a in grow and num(grow[a]["mucin_lift"]) is not None]

        # COMPETITION — best-case: the group isolate with the strongest team result, copied wholesale
        comp_members = [comp[a] for a in members if a in comp]
        best = max(comp_members, key=lambda r: (num(r["best_team_inhib_pa"]) or -1e9,
                                                 num(r["best_solo_inhib_pa"]) or -1e9)) if comp_members else None

        rows.append({
            "strain_group": grp, "representative_asma_id": s["representative_asma_id"],
            "genus": genus, "species": species, "n_isolates": s.get("n_isolates"),
            "is_candidate": is_cand, "candidate_review": review, "bsl_level": bsl,
            "hemolysis_beta": beta, "hemolysis_concern": (beta == "Y") if beta is not None else None,
            "amr_resistance_count_prov": max(res) if res else None,
            "amr_gene_count": max(genes) if genes else None,
            "grows_scfm": grows, "scfm_od": max(scfm_ods) if scfm_ods else None,
            "mucin_lift": max(lifts) if lifts else None,
            "comp_best_solo_pa": best["best_solo_inhib_pa"] if best else None,
            "comp_best_team_pa": best["best_team_inhib_pa"] if best else None,
            "comp_best_partner": best["best_partner"] if best else None,
            "comp_synergy_pa": best["suppressive_synergy_pa"] if best else None,
            "comp_n_formulations": best["n_formulations_pa"] if best else None,
            "tissue": None, "mouse": None, "airway_ubiquity": None,
            "decision": None, "decision_reason": None,
        })

    rows.sort(key=lambda r: (0 if r["is_candidate"] else 1,
                             -(num(r["comp_best_team_pa"]) if num(r["comp_best_team_pa"]) is not None else -1e9),
                             -int(r["n_isolates"] or 0)))
    write_table(rows, GOLD_COLS, GOLD)
    write_xlsx(rows, GOLD_COLS, GOLD + ".xlsx")

    cand = [r for r in rows if r["is_candidate"]]
    unreviewed = sum(1 for r in cand if r["candidate_review"] == "unreviewed")
    print(f"gold_unified_sheet -> {len(rows)} strains ({len(cand)} candidate, {len(rows)-len(cand)} pathogen/opportunistic)")
    print(f"    candidates flagged for safety review : {sum(1 for r in cand if r['candidate_review']=='review')} genus-watchlist + {unreviewed} unreviewed")
    print(f"    candidates that grow in SCFM         : {sum(1 for r in cand if r['grows_scfm']=='Y')}")
    print(f"    candidates with competition data     : {sum(1 for r in cand if r['comp_best_team_pa'] not in (None,''))}")
    print(f"    -> ../data/gold/gold_unified_sheet.{{csv,parquet,xlsx}}")


def write_xlsx(rows, columns, path):
    import openpyxl
    from openpyxl.styles import Font
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "unified_sheet (PRELIM)"
    ws.append(columns)
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in rows:
        ws.append([_x(r.get(c)) for c in columns])
    ws.freeze_panes = "A2"
    for i, c in enumerate(columns, 1):
        width = 22 if c in ("species", "comp_best_partner", "decision_reason") else 13
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # "_about" sheet — tells scientists they are in control
    ab = wb.create_sheet("_about")
    for line in _about_lines():
        ab.append([line])
    ab.column_dimensions["A"].width = 110
    ab["A1"].font = Font(bold=True, size=13)
    wb.save(path)


def _about_lines():
    c = CFG
    return [
        "PROTECT — Unified Formulation Decision Sheet (PRELIMINARY DRAFT)",
        "",
        "This is a living draft built by the data team (Spencer + Alex). It WILL change with your feedback",
        "and as new data arrives. Nothing here is a final decision.",
        "",
        "YOU (the biologists) are in control of every threshold and cutoff. These are our best-guess defaults;",
        "tell us what you want and we recompute the whole card in minutes. Current settings:",
        f"    - AMR 'resistant' cutoff      : >= {c['safety']['amr_measured']['resistance_cutoff_pct']}% growth under the drug",
        f"    - Grows-in-SCFM cutoff        : >= {c['viability']['scfm_grow_min_od']} OD600 (background-subtracted)",
        f"    - Competition replicate value : {c['competition']['replicate_aggregation']} (of repeat wells)",
        f"    - Competition report bar      : >= {c['competition']['report_bar_pct']}% knock-down of PA",
        f"    - Strain rollup              : safety = {c['gold']['safety_aggregation']}, competition = {c['gold']['competition_aggregation']}",
        "",
        "Safety / candidacy: 'is_candidate = True' means 'NOT a known pathogen' — it is NOT a safety clearance.",
        "The candidate list comes from data/reference/species_safety.csv, which is a TEAM-OWNED interim list",
        "meant to be replaced by Gwyn's BSL-1 list. 'candidate_review = review/unreviewed' flags strains a",
        "biologist should still vet.",
        "",
        "Blank cells mean 'not screened yet', not 'no result'. Tissue / mouse / ubiquity columns fill in later.",
        "Questions or changes: Spencer.",
    ]


def _x(v):
    if v is None:
        return ""
    if isinstance(v, str):
        try:
            return float(v) if v.strip() != "" else v
        except ValueError:
            return v
    return v


if __name__ == "__main__":
    main()
