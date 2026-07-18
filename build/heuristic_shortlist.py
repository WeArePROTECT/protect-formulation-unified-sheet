#!/usr/bin/env python3
"""
heuristic_shortlist.py -- the config-driven formulation decision engine.

Reads the gold decision card and the TEAM-OWNED switchboard
(config/formulation_criteria.yaml), applies whatever gates + ranking the
switchboard says, and writes a ranked candidate shortlist. This is Adam's
"pass/fail formulae that allow ranking": must-pass gates thin the field, rank
criteria order the survivors, and every choice is a switch in the YAML, not code.

The engine is intentionally DUMB: it knows nothing about biology. It only knows
how to apply gate / rank / off + a missing-data policy to columns. All the
judgment lives in the config, which the biologists own. Change the config,
re-run, done.

STRAWMAN DEFAULTS ONLY -- every threshold/mode is provisional and team-owned.
Decisions: docs/decisions/heuristic_shortlist_decisions.md

Run:  python heuristic_shortlist.py
Out:  ../data/gold/formulation_shortlist.{csv,parquet,xlsx}
"""
import os
import re
from lib_ids import read_delimited, write_table

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
GOLD_CARD = os.path.join(ROOT, "data", "gold", "gold_unified_sheet.csv")
CRITERIA_YAML = os.path.join(ROOT, "config", "formulation_criteria.yaml")
OUT = os.path.join(ROOT, "data", "gold", "formulation_shortlist")

_CMP = {
    "==": lambda a, b: a == b, "!=": lambda a, b: a != b,
    ">=": lambda a, b: a >= b, "<=": lambda a, b: a <= b,
    ">": lambda a, b: a > b, "<": lambda a, b: a < b,
}
_TRUEISH = {"y", "yes", "true", "1"}
_FALSEISH = {"n", "no", "false", "0"}


# ---------------------------------------------------------------- small helpers
def load_criteria():
    import yaml
    with open(CRITERIA_YAML) as fh:
        return yaml.safe_load(fh)


def is_blank(v):
    return v is None or (isinstance(v, str) and v.strip() == "")


def as_number(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _int(v):
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return 0


def fmt(v):
    n = as_number(v)
    return str(v) if n is None else f"{n:g}"


def cell_with_source(row, crit):
    """Effective (value, source_column), applying column_fallback when primary is blank."""
    v = row.get(crit["column"])
    if not is_blank(v):
        return v, crit["column"]
    fb = crit.get("column_fallback")
    if fb and not is_blank(row.get(fb)):
        return row.get(fb), fb
    return None, None


def rank_value(v):
    """Cell -> sortable float (numbers as-is; Y/N -> 1/0); None if not rankable."""
    if v is None:
        return None
    n = as_number(v)
    if n is not None:
        return n
    s = str(v).strip().lower()
    if s in _TRUEISH:
        return 1.0
    if s in _FALSEISH:
        return 0.0
    return None


def parse_test(pass_when):
    """'<= 2' -> ('<=', 2.0);  '== N' -> ('==', 'N').  Numeric operand if it parses as float."""
    m = re.match(r"\s*(==|!=|>=|<=|>|<)\s*(.+?)\s*$", str(pass_when))
    if not m:
        raise ValueError(f"bad pass_when {pass_when!r} (expected e.g. '== N' or '<= 2')")
    op, rhs = m.group(1), m.group(2)
    n = as_number(rhs)
    return op, (n if n is not None else rhs)


def gate_result(value, pass_when):
    """-> True / False, or None if a present value can't be compared to the test."""
    op, operand = parse_test(pass_when)
    if isinstance(operand, float):
        v = as_number(value)
        return None if v is None else _CMP[op](v, operand)
    return _CMP[op](str(value).strip(), str(operand))


# ------------------------------------------------------------------- evaluation
def evaluate(rows, cfg):
    criteria = cfg.get("criteria", [])
    gates = [c for c in criteria if c.get("mode") == "gate"]
    ranks = [c for c in criteria if c.get("mode") == "rank"]
    rank_by_name = {c["name"]: c for c in ranks}
    # Be lenient: a name in ranking.order that isn't a rank-mode criterion right now
    # (e.g. the scientist just flipped it to a gate) is skipped with a note, not an
    # error -- so flipping ONE mode: field never breaks the build.
    order = []
    for nm in cfg.get("ranking", {}).get("order", []):
        if nm in rank_by_name:
            order.append(nm)
        else:
            print(f"  note: ranking.order lists '{nm}', which is not a rank-mode criterion "
                  f"right now -> skipping it for ranking (change its mode to 'rank' to use it)")
    # rank criteria not named in `order` are still shown, appended after the ordered ones
    shown_ranks = order + [c["name"] for c in ranks if c["name"] not in order]
    unscreened_tier = bool(cfg.get("ranking", {}).get("unscreened_tier", False))
    primary = order[0] if order else None

    out = []
    for row in rows:
        gate_status, failed = {}, []
        for g in gates:
            val, _ = cell_with_source(row, g)
            miss = (g.get("missing") or "pass").lower()
            if val is None:
                gate_status[g["name"]] = "untested"
                if miss == "fail":
                    failed.append(g["name"])
                continue
            res = gate_result(val, g["pass_when"])
            if res is None:                       # present but incomparable -> honor missing policy
                gate_status[g["name"]] = "untested"
                if miss == "fail":
                    failed.append(g["name"])
            elif res:
                gate_status[g["name"]] = "pass"
            else:
                gate_status[g["name"]] = "FAIL"
                failed.append(g["name"])
        passes = not failed

        disp, keys, src = {}, {}, {}
        for nm in shown_ranks:
            val, source = cell_with_source(row, rank_by_name[nm])
            disp[nm], src[nm], keys[nm] = val, source, rank_value(val)
        unscreened = bool(passes and primary is not None and keys.get(primary) is None)

        out.append({
            "row": row, "passes": passes, "failed": failed, "gate_status": gate_status,
            "disp": disp, "keys": keys, "src": src, "unscreened": unscreened,
        })
    return out, gates, ranks, order, shown_ranks, rank_by_name, unscreened_tier, primary


def sort_key(ev, order, rank_by_name):
    """Lexicographic key (smaller = better) honoring higher_is_better + missing policy."""
    parts = []
    for nm in order:
        c = rank_by_name[nm]
        rv = ev["keys"].get(nm)
        higher = c.get("higher_is_better", True)
        miss = (c.get("missing") or "rank_last").lower()
        if rv is None:
            parts.append((0, 0.0) if miss == "rank_first" else (1, 0.0))
        else:
            keyval = -rv if higher else rv
            parts.append((1, keyval) if miss == "rank_first" else (0, keyval))
    return tuple(parts)


# ----------------------------------------------------------------------- output
GROUP_STATUS = {0: "shortlist", 1: "shortlist_unscreened", 2: "excluded"}


def make_reason(ev, order, primary):
    if not ev["passes"]:
        return "excluded: failed " + ", ".join(ev["failed"])
    head = (f"passes gates; rank #{ev['position']}" if not ev["unscreened"]
            else f"passes gates; not yet screened for {primary}")
    bits = [head] + [f"{nm}=" + ("n/a" if ev["disp"].get(nm) is None else fmt(ev["disp"][nm]))
                     for nm in order]
    return "; ".join(bits)


def select_pool(rows, cfg):
    """Apply the candidates_only switch -- the row filter main() feeds into the engine."""
    if cfg.get("candidates_only", True):
        return [r for r in rows if str(r.get("is_candidate")).strip().lower() == "true"]
    return list(rows)


def assemble(rows, cfg):
    """PURE logic: rows + config -> (out_rows, out_cols, evals, gates, order, primary). No file IO.

    Kept IO-free on purpose so the test suite can drive it on synthetic data with a known
    right answer (tests/test_golden.py). main() is just select_pool -> assemble -> write.
    """
    evals, gates, ranks, order, shown_ranks, rank_by_name, unscreened_tier, primary = evaluate(rows, cfg)

    def group(ev):
        if not ev["passes"]:
            return 2
        if unscreened_tier and ev["unscreened"]:
            return 1
        return 0

    for ev in evals:
        ev["group"] = group(ev)
        ev["skey"] = sort_key(ev, order, rank_by_name)
    evals.sort(key=lambda e: (e["group"], e["skey"], -_int(e["row"].get("n_isolates"))))

    pos = 0
    for ev in evals:
        if ev["group"] == 0:
            pos += 1
            ev["position"] = pos
        else:
            ev["position"] = None

    gate_names = [g["name"] for g in gates]
    base = ["rank", "status", "strain_group", "representative_asma_id", "genus", "species", "n_isolates"]
    gcols = [f"gate_{n}" for n in gate_names]
    vcols = []
    for n in shown_ranks:
        vcols.append(f"value_{n}")
        if rank_by_name[n].get("column_fallback"):
            vcols.append(f"source_{n}")
    out_cols = base + gcols + vcols + ["gates_failed", "shortlist_reason"]

    out_rows = []
    for ev in evals:
        r = ev["row"]
        rec = {
            "rank": ev["position"], "status": GROUP_STATUS[ev["group"]],
            "strain_group": r.get("strain_group"), "representative_asma_id": r.get("representative_asma_id"),
            "genus": r.get("genus"), "species": r.get("species"), "n_isolates": r.get("n_isolates"),
            "gates_failed": "; ".join(ev["failed"]),
            "shortlist_reason": make_reason(ev, order, primary),
        }
        for n in gate_names:
            rec[f"gate_{n}"] = ev["gate_status"].get(n, "")
        for n in shown_ranks:
            v = ev["disp"].get(n)
            rec[f"value_{n}"] = "" if v is None else v
            if rank_by_name[n].get("column_fallback"):
                rec[f"source_{n}"] = ev["src"].get(n) or ""
        out_rows.append(rec)
    return out_rows, out_cols, evals, gates, order, primary


def main():
    cfg = load_criteria()
    rows = select_pool(list(read_delimited(GOLD_CARD, ",")), cfg)
    out_rows, out_cols, evals, gates, order, primary = assemble(rows, cfg)
    write_table(out_rows, out_cols, OUT)
    write_xlsx(out_rows, out_cols, OUT + ".xlsx", cfg, evals, gates, order, primary)
    print_summary(evals, gates, order, primary)


# --------------------------------------------------------------- xlsx + summary
def print_summary(evals, gates, order, primary):
    n = len(evals)
    npass = sum(1 for e in evals if e["passes"])
    nrank = sum(1 for e in evals if e["group"] == 0)
    nunscr = sum(1 for e in evals if e["group"] == 1)
    nexcl = sum(1 for e in evals if e["group"] == 2)
    print(f"formulation_shortlist -> {n} candidate strains evaluated")
    for g in gates:
        f = sum(1 for e in evals if g["name"] in e["failed"])
        print(f"    gate '{g['name']}' ({g['column']} {g['pass_when']}): {f} fail")
    print(f"    pass ALL gates                          : {npass}")
    print(f"    -> ranked shortlist ('{primary}' present): {nrank}")
    print(f"    -> pass gates but '{primary}' unscreened : {nunscr}")
    print(f"    excluded (failed >=1 gate)              : {nexcl}")
    print(f"    -> ../data/gold/formulation_shortlist.{{csv,parquet,xlsx}}")


def switchboard_lines(cfg, evals, gates, order, primary):
    L = [
        "PROTECT -- Formulation Shortlist (PRELIMINARY, strawman defaults)",
        "",
        "This shortlist is produced by build/heuristic_shortlist.py from the TEAM-OWNED switchboard",
        "config/formulation_criteria.yaml. Every gate, ranking, and cutoff below is a switch YOU control:",
        "edit the YAML and re-run  bash build/run_all.sh  to recompute the whole shortlist.",
        "",
        f"candidates_only: {cfg.get('candidates_only', True)}",
        "",
        "CRITERIA (the switches currently set):",
    ]
    for c in cfg.get("criteria", []):
        mode = c.get("mode")
        if mode == "gate":
            L.append(f"    [GATE] {c['name']}: must pass  {c['column']} {c.get('pass_when')}"
                     f"   (missing -> {c.get('missing', 'pass')})")
        elif mode == "rank":
            hib = c.get("higher_is_better", True)
            fb = f", fallback {c['column_fallback']}" if c.get("column_fallback") else ""
            L.append(f"    [RANK] {c['name']}: order by {c['column']}{fb}"
                     f" ({'higher' if hib else 'lower'} is better; missing -> {c.get('missing', 'rank_last')})")
        else:
            L.append(f"    [OFF ] {c['name']}: ignored")
    npass = sum(1 for e in evals if e["passes"])
    nrank = sum(1 for e in evals if e["group"] == 0)
    nunscr = sum(1 for e in evals if e["group"] == 1)
    nexcl = sum(1 for e in evals if e["group"] == 2)
    L += [
        "",
        f"ranking order (primary first): {', '.join(order) if order else '(none)'}",
        f"unscreened_tier: {cfg.get('ranking', {}).get('unscreened_tier', False)}"
        f"  (strains with no '{primary}' data are listed separately, not ranked as losers)",
        "",
        "RESULT with the current switches:",
    ]
    for g in gates:
        f = sum(1 for e in evals if g["name"] in e["failed"])
        L.append(f"    gate '{g['name']}' failed by {f} strain(s)")
    L += [
        f"    {npass} pass all gates -> {nrank} ranked + {nunscr} pass-but-unscreened; {nexcl} excluded.",
        "",
        "STATUS = PRELIMINARY. Wet-lab inputs are pre-QC (Sun-Young). Blank cells mean 'not screened yet',",
        "not 'no result'. is_candidate=True means 'not a known pathogen', NOT 'safety cleared'.",
        "You own every switch. Tell Spencer what to change, or edit config/formulation_criteria.yaml.",
    ]
    return L


def _x(v):
    if v is None:
        return ""
    if isinstance(v, str):
        try:
            return float(v) if v.strip() != "" else v
        except ValueError:
            return v
    return v


def write_xlsx(rows, columns, path, cfg, evals, gates, order, primary):
    import openpyxl
    from openpyxl.styles import Font
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "shortlist (PRELIM)"
    ws.append(columns)
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in rows:
        ws.append([_x(r.get(c)) for c in columns])
    ws.freeze_panes = "A2"
    for i, c in enumerate(columns, 1):
        width = 26 if c in ("species", "shortlist_reason") else 14
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    sb = wb.create_sheet("_switchboard")
    for line in switchboard_lines(cfg, evals, gates, order, primary):
        sb.append([line])
    sb.column_dimensions["A"].width = 100
    sb["A1"].font = Font(bold=True, size=13)
    wb.save(path)


if __name__ == "__main__":
    main()
